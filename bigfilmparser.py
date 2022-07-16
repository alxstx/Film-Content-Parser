from selenium import webdriver
import requests
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import sys
import threading
from urllib3 import exceptions
from pathlib import Path
import atexit

input_path = input('path:')
url_list = []
notneccesarylist = ['phrases', 'stories', 'snacks', 'trailers', 'puns', 'facts', 'theories', 'things', 'fans']
thread_list = []
from http import client


def get_html2(url, driver1):
    try:
        html = driver1.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
    except:
        html = ''
        pass
    return html


def get_html(url, params):
    r = requests.get(url, params=params)
    return r


lists_websites = [
    'https://www.ranker.com/list/best-korean-dramas-2019/molly-gander?ref=browse_list&l=1']

all_data = {'Название списка': [], 'Дата публикаций': [], 'Имя автора публикации': [],
            'Описание списка': [], 'Название': [], 'Дата премьеры': [],
            'Порядковый номер в списке': [], 'Url': []}
data = {'Название списка(фильмы)': [], 'Дата публикаций(фильмы)': [], 'Имя автора публикации(фильмы)': [],
        'Описание списка(фильмы)': [], 'Название фильма/сериала': [], 'Дата премьеры': [],
        'Порядковый номер фильма в списке': [], 'Url film': []}
data2 = {'Название списка(персонажи)': [],
         'Дата публикаций(персонажи)': [], 'Имя автора публикации(персонажи)': [],
         'Описание списка(персонажи)': [], 'Имя актёра/персонажа': [],
         'Порядковый номер персонажа в списке': [], 'Url персонаж': []}
data3 = {'Название списка(актёры)': [],
         'Дата публикаций(актёры)': [], 'Имя автора публикации(актёры)': [],
         'Описание списка(актёры)': [], 'Имя актёра/персонажа': [],
         'Порядковый номер актёра в списке': [], 'Url актёр': []}
data4 = {'Название списка(сериалы)': [], 'Дата публикаций(сериалы)': [], 'Имя автора публикации(сериалы)': [],
         'Описание списка(сериалы)': [], 'Название фильма/сериала': [], 'Дата премьеры': [],
         'Порядковый номер сериала в списке': [], 'Url series': []}
linklist = []
a = open('urllist11.txt', 'r')
numberoflinks = len(a.readlines())  # len(a.read())
print(numberoflinks)
a.close()


def read_file():
    a = open('urllist11.txt', 'r')
    link = a.readline()
    linklist.append(link)
    a.close()
    with open("urllist11.txt", "r") as f:
        lines = f.readlines()
    with open("urllist11.txt", "w") as f:
        for line in lines:
            if line not in linklist:
                f.write(line)
    return link


flag2 = True
driver_lists = []


def get_all_data(flag, flag2):
    driver1 = webdriver.Chrome(executable_path=input_path)
    driver_lists.append(driver1)
    if flag:
        for i in range(int(numberoflinks / n)):
            try:
                link = read_file()
                link = link.strip().replace('\n', '')
                print(link)
                if link != '':
                    ddd = func()
                    if ddd == 'stop':
                        end_that()
                    print(link)
                    try:
                       driver1.get(link)
                    except:
                        pass
                    try:
                       height = driver1.execute_script("return document.body.scrollHeight")
                    except:
                        height = 1
                    try:
                       for scrol in range(200, height * 10, 200):
                        if flag2:
                            driver1.execute_script(f"window.scrollTo(0,{scrol})")
                            time.sleep(0.2)
                        try:
                            itemi = driver1.find_element_by_xpath('//*[@id="__next"]/article/div[3]/div/div[2]/a')
                        except:
                            itemi = None
                        if itemi != None:
                            try:
                                driver1.find_element_by_xpath('//*[@id="__next"]/article/div[3]/div/div[2]/a').click()
                                flag2 = False
                            except:
                                WebDriverWait(driver1, 60).until(
                                    EC.invisibility_of_element(
                                        (By.XPATH, '//*[@id="__next"]/article/div[3]/div/div[2]/a')))
                                flag2 = False
                    except:
                        pass
                    try:
                       html = get_html2(link, driver1)
                       soup = BeautifulSoup(html, 'lxml')
                       items = soup.find_all('li')
                       testnumber = soup.find('strong', 'gridItem_rank__3Q_TO')
                    except:
                        testnumber=None
                        soup=None
                        items=[]
                        pass
                    try:
                        title = soup.find('h1', 'title_name__WcUIn').text
                    except:
                        title = ''
                    notneccesarycheck = [s in title for s in notneccesarylist]
                    if items != [] and testnumber != None and True not in notneccesarycheck:
                        try:
                           date = soup.find('div', 'stats_main__2WISH').find_next('span').text
                        except:
                            date = ''
                        print(date)
                        try:
                            namepublicator = soup.find('a', 'byline_name__15EbD byline_authorLink__3RaEf').text
                        except:
                            namepublicator = ''.join(soup.find('div', 'byline_name__15EbD').find_all(text=True))
                        print(namepublicator)
                        try:
                           description = ''.join(soup.find('div', 'description_text__3Xyg2').find_all(text=True))
                        except:
                            description = ''
                        print(description)
                        for item in items:
                            try:
                                number = item.find('strong', 'gridItem_rank__3Q_TO').text
                            except:
                                number = ''
                            print(number)
                            try:
                                driver1.find_element_by_xpath(item.find('svg')).click()
                                html = driver1.page_source
                                soup = BeautifulSoup(html, 'lxml')
                                released = soup.find('div',
                                                     'richText_container__4e_Bv nodePopup_wikiText__12KWm').find_next(
                                    'span').text
                                driver1.find_element_by_xpath(
                                    '/html/body/div[1]/article/div[3]/div/ul/div[2]/div/span/svg/use//svg/path').click()
                            except:
                                released = 'unknown'
                            print(released)

                            try:
                                filmname = item.find('h2').find_next('a').text
                                print(filmname)
                            except:
                                filmname = ''

                                # saving
                            if filmname != '' and number != '':
                                title = title.replace('\xa0', ' ').replace('\n', ' ')
                                if filmname == '\xa0\n      more': filmname = 'No Name'
                                all_data['Название списка'].append(title)
                                all_data['Дата публикаций'].append(date)
                                all_data['Имя автора публикации'].append(namepublicator)
                                all_data['Описание списка'].append(description)
                                all_data['Название'].append(filmname)
                                all_data['Дата премьеры'].append(released)
                                all_data['Порядковый номер в списке'].append(number)
                                all_data['Url'].append(link)
                                if 'movies' or 'films' or 'dramas' in title.lower():
                                    data['Название списка(фильмы)'].append(title)
                                    data['Дата публикаций(фильмы)'].append(date)
                                    data['Имя автора публикации(фильмы)'].append(namepublicator)
                                    data['Описание списка(фильмы)'].append(description)
                                    data['Название фильма/сериала'].append(filmname)
                                    data['Дата премьеры'].append(released)
                                    data['Порядковый номер фильма в списке'].append(number)
                                    data['Url film'].append(link)
                                elif 'shows' or 'series' or 'cartoons' or 'sitcoms' in title.lower():
                                    data4['Название списка(сериалы)'].append(title)
                                    data4['Дата публикаций(сериалы)'].append(date)
                                    data4['Имя автора публикации(сериалы)'].append(namepublicator)
                                    data4['Описание списка(сериалы)'].append(description)
                                    data4['Название фильма/сериала'].append(filmname)
                                    data4['Дата премьеры'].append(released)
                                    data4['Порядковый номер сериала в списке'].append(number)
                                    data4['Url series'].append(link)

                                elif 'villains' or 'characters' or 'figures' in title.lower():
                                    data2['Название списка(персонажи)'].append(title)
                                    data2['Дата публикаций(персонажи)'].append(date)
                                    data2['Имя автора публикации(персонажи)'].append(namepublicator)
                                    data2['Описание списка(персонажи)'].append(description)
                                    data2['Имя актёра/персонажа'].append(filmname)
                                    data2['Порядковый номер персонажа в списке'].append(number)
                                    data2['Url персонаж'].append(link)
                                elif 'actors' or 'actresses' or 'writers':
                                    data3['Название списка(персонажи)'].append(title)
                                    data3['Дата публикаций(персонажи)'].append(date)
                                    data3['Имя автора публикации(персонажи)'].append(namepublicator)
                                    data3['Описание списка(персонажи)'].append(description)
                                    data3['Имя актёра/персонажа'].append(filmname)
                                    data3['Порядковый номер персонажа в списке'].append(number)
                                    data3['Url персонаж'].append(link)
                            else:
                                pass

                    else:
                        print('oops wrong website')
                else:
                    print('letsgo')
            except KeyboardInterrupt or exceptions.MaxRetryError or client.RemoteDisconnected or ConnectionRefusedError or exceptions.NewConnectionError or exceptions.ProtocolError:
                print('endd')



obzhyfile = input('название файла(отсек общий):')
filmfile = input('название файла(отсек фильмы):')
serialfile = input('название файла(отсек сериалы):')
personfile = input('название файла(отсек персоны):')
charactersfile = input('название файла(отсек персонажи):')


def get_from_file():
    df = pd.read_excel(obzhyfile)
    df2 = pd.read_excel(filmfile)
    df3 = pd.read_excel(serialfile)
    df4 = pd.read_excel(personfile)
    df5 = pd.read_excel(charactersfile)
    for all in all_data.keys():
        all_data[all] = df[all].tolist() + all_data[all]
    for dat in data.keys():
        data[dat] = df2[dat].tolist() + data[dat]
    for dat2 in data2.keys():
        data2[dat2] = df5[dat2].tolist() + data2[dat2]
    for dat3 in data3:
        data3[dat3] = df4[dat3].tolist() + data3[dat3]
    for dat4 in data4:
        data4[dat4] = df3[dat4].tolist() + data4[dat4]
    print(all_data, '\n', data, '\n', data2, '\n', data3, '\n', data4)
    wb1 = load_workbook(obzhyfile)
    sheet1 = wb1['Лист1']
    wb2 = load_workbook(filmfile)
    sheet2 = wb2['Лист1']
    wb3 = load_workbook(serialfile)
    sheet3 = wb3['Лист1']
    wb4 = load_workbook(personfile)
    sheet4 = wb4['Лист1']
    wb5 = load_workbook(charactersfile)
    sheet5 = wb5['Лист1']
    sheet1.delete_rows(0, sheet1.max_row)
    sheet2.delete_rows(1, sheet2.max_row)
    sheet3.delete_rows(2, sheet3.max_row)
    sheet4.delete_rows(3, sheet4.max_row)
    sheet5.delete_rows(4, sheet5.max_row)
    wb1.save(obzhyfile)
    wb2.save(filmfile)
    wb3.save(serialfile)
    wb4.save(personfile)
    wb5.save(charactersfile)


lss = []


def save_data():
    print('saving...')
    workbook = xlsxwriter.Workbook(obzhyfile)
    workbook2 = xlsxwriter.Workbook(filmfile)
    workbook3 = xlsxwriter.Workbook(serialfile)
    workbook4 = xlsxwriter.Workbook(personfile)
    workbook5 = xlsxwriter.Workbook(charactersfile)
    obthee = workbook.add_worksheet('Воркшит намбер 1')
    filmi = workbook2.add_worksheet('Воркшит намбер 1')
    serial = workbook3.add_worksheet('Воркшит намбер 1')
    person = workbook4.add_worksheet('Воркшит намбер 1')
    characteri = workbook5.add_worksheet('Воркшит намбер 1')
    print('saving...')
    lss.extend([workbook, workbook2, workbook3, workbook4, workbook5, obthee, filmi, serial, person, characteri])


def save_data2():
    workbook, workbook2, workbook3, workbook4, workbook5, obthee, filmi, serial, person, characteri = lss
    print('woking?')
    for i, ob in enumerate(all_data.keys()):
        obthee.write(0, i, ob)
    for g, obt in enumerate(all_data.values()):
        for f, o in enumerate(obt):
            obthee.write(f + 1, g, o)
    print('error?')
    workbook.close()
    print('first')
    for i1, ob1 in enumerate(data.keys()):
        filmi.write(0, i1, ob1)
    for g1, obt1 in enumerate(data.values()):
        if obt1 != []:
            for f1, o1 in enumerate(obt1):
                filmi.write(f1 + 1, g1, o1)
    workbook2.close()
    print('second')
    for i2, ob2 in enumerate(data2.keys()):
        characteri.write(0, i2, ob2)
    for g2, obt2 in enumerate(data2.values()):
        if obt2 != []:
            for f2, o2 in enumerate(obt2):
                characteri.write(f2 + 1, g2, o2)
    workbook3.close()
    print('third')
    for i3, ob3 in enumerate(data3.keys()):
        person.write(0, i3, ob3)
    for g3, obt3 in enumerate(data3.values()):
        if obt3 != []:
            for f3, o3 in enumerate(obt3):
                person.write(f3 + 1, g3, o3)
    workbook4.close()
    print('fourth')
    for i4, ob4 in enumerate(data4.keys()):
        serial.write(0, i4, ob4)
    for g4, obt4 in enumerate(data4.values()):
        if obt4 != []:
            for f4, o4 in enumerate(obt4):
                serial.write(f4 + 1, g4, o4)
    workbook5.close()
    print('fifth')
    print('SAVED')
    print('NEARLY SAVED')
    print('END SAVED')


def end_that():
    print('end soon')
    if Path(obzhyfile).exists():
        try:
            get_from_file()
        except FileNotFoundError:
            pass
    else:
        print('file doesnt exist')
        save_data()
        save_data2()
        print('THE END')



flag = True
# atexit.register(end_that)
n = int(input('количество окон:'))
if __name__ == '__main__':
    try:
        for i in range(n):
            print(n)
            t = threading.Thread(name='Test {}'.format(i), target=get_all_data, args=(flag, flag2,))
            t.daemon = True
            t.start()
            time.sleep(1)
            thread_list.append(t)
        for thread in thread_list:
            thread.join()
    except KeyboardInterrupt or exceptions.MaxRetryError or client.RemoteDisconnected or ConnectionRefusedError or exceptions.NewConnectionError or exceptions.ProtocolError:
        print()
