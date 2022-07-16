import pandas as pd
from openpyxl import load_workbook
import xlsxwriter

lss = []
all_data = {'Название списка': [], 'Дата публикаций': [], 'Имя автора публикации': [],
            'Описание списка': [], 'Название': [], 'Дата премьеры': [],
            'Порядковый номер в списке': [], 'Url': []}
data = {'Название списка(фильмы)': [], 'Дата публикаций(фильмы)': [], 'Имя автора публикации(фильмы)': [],
        'Описание списка(фильмы)': [], 'Название фильма/сериала': [], 'Дата премьеры': [],
        'Порядковый номер фильма в списке': [], 'Url film': []}
data2 = {'Название списка(персонажи)': [],
         'Дата публикаций(персонажи)': [], 'Имя автора публикации(персонажи)': [],
         'Описание списка(персонажи)': [], 'Имя актёра/персонажа': [],
         'Порядковый номер персонажа в списке': [], 'Url персонаж': []}
data3 = {'Название списка(актёры)': [],
         'Дата публикаций(актёры)': [], 'Имя автора публикации(актёры)': [],
         'Описание списка(актёры)': [], 'Имя актёра/персонажа': [],
         'Порядковый номер актёра в списке': [], 'Url актёр': []}
data4 = {'Название списка(сериалы)': [], 'Дата публикаций(сериалы)': [], 'Имя автора публикации(сериалы)': [],
         'Описание списка(сериалы)': [], 'Название фильма/сериала': [], 'Дата премьеры': [],
         'Порядковый номер сериала в списке': [], 'Url series': []}

obzhyfile = input('название файла(отсек общий):')
filmfile = input('название файла(отсек фильмы):')
serialfile = input('название файла(отсек сериалы):')
personfile = input('название файла(отсек персоны):')
charactersfile = input('название файла(отсек персонажи):')


def save_data():
    workbook = xlsxwriter.Workbook(obzhyfile)
    workbook2 = xlsxwriter.Workbook(filmfile)
    workbook3 = xlsxwriter.Workbook(serialfile)
    workbook4 = xlsxwriter.Workbook(personfile)
    workbook5 = xlsxwriter.Workbook(charactersfile)
    obthee = workbook.add_worksheet('Воркшит намбер 1')
    filmi = workbook2.add_worksheet('Воркшит намбер 1')
    serial = workbook3.add_worksheet('Воркшит намбер 1')
    person = workbook4.add_worksheet('Воркшит намбер 1')
    characteri = workbook5.add_worksheet('Воркшит намбер 1')
    lss.extend([workbook, workbook2, workbook3, workbook4, workbook5, obthee, filmi, serial, person, characteri])


def save_data2():
    workbook, workbook2, workbook3, workbook4, workbook5, obthee, filmi, serial, person, characteri = lss
    for i, ob in enumerate(all_data.keys()):
        obthee.write(0, i, ob)
    for i1, ob1 in enumerate(data.keys()):
        filmi.write(0, i1, ob1)
    for i2, ob2 in enumerate(data2.keys()):
        characteri.write(0, i2, ob2)
    for i3, ob3 in enumerate(data3.keys()):
        person.write(0, i3, ob3)
    for i4, ob4 in enumerate(data4.keys()):
            serial.write(0, i4, ob4)
    workbook.close()
    workbook2.close()
    workbook3.close()
    workbook4.close()
    workbook5.close()


def get_from_file():
    df = pd.read_excel(obzhyfile)
    df2 = pd.read_excel(filmfile)
   # df3 = pd.read_excel(serialfile)
    df4 = pd.read_excel(personfile)
    df5 = pd.read_excel(charactersfile)
    for all in all_data.keys():
        all_data[all] = df[all].tolist() + all_data[all]
    for dat in data.keys():
        data[dat] = df2[dat].tolist() + data[dat]
    for dat2 in data2.keys():
        data2[dat2] = df5[dat2].tolist() + data2[dat2]
    for dat3 in data3:
        data3[dat3] = df4[dat3].tolist() + data3[dat3]
    #for dat4 in data4:
       # data4[dat4] = df3[dat4].tolist() + data4[dat4]
    wb1 = load_workbook(obzhyfile)
    sheet1 = wb1['Воркшит намбер 1']
    wb2 = load_workbook(filmfile)
    sheet2 = wb2['Воркшит намбер 1']
   # wb3 = load_workbook(serialfile)
  #  sheet3 = wb3['Лист1']
    wb4 = load_workbook(personfile)
    sheet4 = wb4['Воркшит намбер 1']
    wb5 = load_workbook(charactersfile)
    sheet5 = wb5['Воркшит намбер 1']
    sheet1.delete_rows(1, sheet1.max_row)
    sheet2.delete_rows(1, sheet2.max_row)
  #  sheet3.delete_rows(2, sheet3.max_row)
    sheet4.delete_rows(1, sheet4.max_row)
    sheet5.delete_rows(1, sheet5.max_row)
    wb1.save(obzhyfile)
    wb2.save(filmfile)
  #  wb3.save(serialfile)
    wb4.save(personfile)
    wb5.save(charactersfile)


get_from_file()
print(all_data)
print(data)
print(data2)
print(data3)
print(data4)
